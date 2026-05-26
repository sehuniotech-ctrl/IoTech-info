from __future__ import annotations

import json
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name == "scripts" else Path("github_wbs_repo")
OUTPUT = REPO_ROOT / "files" / "WBS_업무정리_순수WBS.xlsx"
ARCHIVE = REPO_ROOT / "archive" / "2026-05-26_WBS_업무정리_순수WBS.xlsx"
DATA = REPO_ROOT / "data" / "wbs_items.json"


def d(value: str) -> date:
    y, m, day = map(int, value.split("-"))
    return date(y, m, day)


rows = [
    ("아이오테크 SEMS SL 개발", "요구사항/범위", "통합진행문서 정리", "미진행", "2026-05-26", "2026-05-28", "01_요구사항/01_통합진행문서_SmartLoad.md 기준으로 현재 목표와 범위 정리", "https://www.notion.so/36c5c9ca0fd38112b54fd55eae428b11"),
    ("아이오테크 SEMS SL 개발", "요구사항/범위", "기능정의 및 구현범위 정리", "미진행", "2026-05-26", "2026-05-29", "02_기능정의_및_구현범위.md 기준으로 구현 범위와 제외 범위 정리", "https://www.notion.so/36c5c9ca0fd3814caea1c55aa24a337f"),
    ("아이오테크 SEMS SL 개발", "요구사항/범위", "지선차단기 다이어그램 정리", "미진행", "2026-05-27", "2026-05-29", "03_지선차단기_다이어그램_정리.md를 기준으로 구성요소와 연결 흐름 정리", "https://www.notion.so/36c5c9ca0fd3811e95e8edac44992d4d"),
    ("아이오테크 SEMS SL 개발", "HW 개발", "칩선정 및 구매성 검토", "미진행", "2026-05-29", "2026-06-03", "01_칩선정_및_구매성.md와 ATM90E26 관련 문서 기준으로 선정 근거 정리", "https://www.notion.so/36c5c9ca0fd38177955cc7732bf59a37"),
    ("아이오테크 SEMS SL 개발", "HW 개발", "HW팀 전달 기능/핀 기준", "미진행", "2026-06-03", "2026-06-07", "HW팀 전달용 기능/핀 기준 최신본 정리", "https://www.notion.so/36c5c9ca0fd3810886a4f106f97ab6fc"),
    ("아이오테크 SEMS SL 개발", "HW 개발", "ATM90E26 기능/핀 정리", "미진행", "2026-06-04", "2026-06-10", "ATM90E26 기능검토/핀사용정리 문서를 기준으로 FW-HW 연결 기준 정리", "https://www.notion.so/36c5c9ca0fd38198ae55fe255625ccb2"),
    ("아이오테크 SEMS SL 개발", "HW 개발", "보조저전력 출력 분리 회로 정리", "미진행", "2026-06-08", "2026-06-14", "보조저전력 출력 분리 블록도/개념회로도/준회로도 자료 정리", "https://www.notion.so/36c5c9ca0fd3811d85f2d310d095c80f"),
    ("아이오테크 SEMS SL 개발", "HW 개발", "회로도", "미진행", "2026-06-10", "2026-06-17", "최신 회로도 파일 위치 확인", "https://www.notion.so/36c5c9ca0fd381d985f1ce78fe292504"),
    ("아이오테크 SEMS SL 개발", "HW 개발", "PCB", "미진행", "2026-06-15", "2026-06-24", "PCB 관련 자료와 진행 상태 확인", "https://www.notion.so/36c5c9ca0fd381e6b45df9803e9e36c8"),
    ("아이오테크 SEMS SL 개발", "HW 개발", "아트웍", "미진행", "2026-06-22", "2026-07-03", "아트웍 진행 파일과 수정사항 확인", "https://www.notion.so/36c5c9ca0fd381b09e0cc1f1a7485102"),
    ("아이오테크 SEMS SL 개발", "FW 개발", "펌웨어 개발 기준 정리", "미진행", "2026-05-29", "2026-06-03", "01_펌웨어개발기준.md와 smartload_fw_current README 기준으로 개발 기준 정리", "https://www.notion.so/36c5c9ca0fd38194b445fe2abe16412a"),
    ("아이오테크 SEMS SL 개발", "FW 개발", "LCU보드 데모 준비", "미진행", "2026-06-03", "2026-06-14", "LCU보드 데모에 필요한 기능과 현재 자료 위치 확인", "https://www.notion.so/36c5c9ca0fd3811db6facdf61c7f349e"),
    ("아이오테크 SEMS SL 개발", "FW 개발", "F107 LCU 타깃 Bring-up", "미진행", "2026-06-10", "2026-06-21", "target_f107의 bring-up 체크리스트와 소스 구조 확인", "https://www.notion.so/36c5c9ca0fd381d8bb54eca9fb78aec5"),
    ("아이오테크 SEMS SL 개발", "FW 개발", "PC 데모/프로토콜 도구 정리", "미진행", "2026-06-10", "2026-06-18", "pc_demo 폴더의 GUI/CLI/시뮬레이터 도구 역할 정리", "https://www.notion.so/36c5c9ca0fd381a69582c1a7e05e0832"),
    ("아이오테크 SEMS SL 개발", "FW 개발", "ATM90E26 드라이버/계측 연동", "미진행", "2026-06-17", "2026-06-28", "atm90e26 관련 src/inc 파일과 transport 구현 확인", "https://www.notion.so/36c5c9ca0fd38104b8eff6ba9602338f"),
    ("아이오테크 SEMS SL 개발", "통신규격", "DCU 통신프로토콜 명세서 정리", "미진행", "2026-06-03", "2026-06-10", "03_지선차단기_DCU_통신프로토콜_명세서.md 기준으로 최종 프로토콜 항목 정리", "https://www.notion.so/36c5c9ca0fd38119bcdeeeed9958d30e"),
    ("아이오테크 SEMS SL 개발", "통신규격", "예시패킷/데이터스케일 정리", "미진행", "2026-06-10", "2026-06-16", "예시패킷과 데이터스케일 추천표를 기준으로 구현/테스트 기준 정리", "https://www.notion.so/36c5c9ca0fd3819ebf3af9e407a26eb3"),
    ("아이오테크 SEMS SL 개발", "통합/데모", "FW-HW 통합 테스트", "미진행", "2026-06-24", "2026-07-05", "데모 성공 기준과 테스트 항목 정리", "https://www.notion.so/36c5c9ca0fd381bcaddfe6762f1a4c84"),
    ("아이오테크 SEMS SL 개발", "회의/보고", "최신 회의메모 정리", "미진행", "2026-05-26", "2026-05-28", "07_회의메모/01_최신회의메모.md 기준으로 결정사항과 후속작업 분리", "https://www.notion.so/36c5c9ca0fd38190a857eba73f8f3ecf"),
    ("아이오테크 SEMS SL 개발", "회의/보고", "최신 보고자료 정리", "미진행", "2026-05-29", "2026-06-05", "09_보고서/01_최신보고자료.md와 주요 PPT/DOCX 산출물 기준으로 최신본 정리", "https://www.notion.so/36c5c9ca0fd3815ebc85fd0b92e71d68"),
    ("아이오테크 SEMS SL 개발", "일정/WBS", "진행률 기준 정리", "미진행", "2026-05-26", "2026-05-29", "06_일정WBS의 진행률 기준과 전체업무 비중 문서를 현재 WBS에 반영", "https://www.notion.so/36c5c9ca0fd381679aefc6301bfda0bc"),
    ("아이오테크 SEMS SL 개발", "단발 요청", "임시 요청/확인사항", "미진행", "2026-05-26", "2026-06-30", "들어오는 요청은 여기 먼저 적고 나중에 제자리로 이동", "https://www.notion.so/36c5c9ca0fd3817fba1adbaf6b8880a1"),
    ("위웨이크 앱 개발", "운동 앱", "기능/운영 정리", "미진행", "2026-05-26", "2026-06-12", "운동 앱의 현재 기능과 수정사항 정리", "https://www.notion.so/36c5c9ca0fd3812ba480d43f622e10a5"),
    ("위웨이크 앱 개발", "정부정책 앱", "기능/운영 정리", "미진행", "2026-05-26", "2026-06-12", "정부정책 앱의 현재 기능과 수정사항 정리", "https://www.notion.so/36c5c9ca0fd381b28a3ecd5ca9833abf"),
    ("개인 관리", "건강", "해야 할 일 정리", "미진행", "2026-05-26", "2026-06-02", "생각나는 일을 먼저 적기", "https://www.notion.so/36c5c9ca0fd381cd9214e2e65c659fee"),
    ("개인 관리", "돈/생활", "해야 할 일 정리", "미진행", "2026-05-26", "2026-06-02", "생각나는 일을 먼저 적기", "https://www.notion.so/36c5c9ca0fd38189b6b4dbe582dc9937"),
]


def merge_runs(sheet, col: int, start_row: int, values: list[object]) -> None:
    run_start = 0
    for idx in range(1, len(values) + 1):
        if idx == len(values) or values[idx] != values[run_start]:
            if idx - run_start > 1:
                sheet.merge_cells(
                    start_row=start_row + run_start,
                    start_column=col,
                    end_row=start_row + idx - 1,
                    end_column=col,
                )
                sheet.cell(start_row + run_start, col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            run_start = idx


def build_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS목록"
    ws.sheet_view.showGridLines = False

    headers = ["1Depth", "2Depth", "3Depth", "상태", "시작일", "종료일", "소요일", "다음 행동", "Notion"]
    ws.append(headers)
    for item in rows:
        ws.append([item[0], item[1], item[2], item[3], d(item[4]), d(item[5]), None, item[6], "열기"])

    for row in range(2, len(rows) + 2):
        ws.cell(row, 7).value = f'=IF(OR(E{row}="",F{row}=""),"",F{row}-E{row}+1)'
        ws.cell(row, 9).hyperlink = rows[row - 2][7]
        ws.cell(row, 9).style = "Hyperlink"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    light_fill = PatternFill("solid", fgColor="EAF3F8")
    stripe_fill = PatternFill("solid", fgColor="F7FBFD")
    thin_gray = Side(style="thin", color="7F7F7F")
    blue_line = Side(style="thin", color="9CCFE5")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1, min_col=1, max_col=9):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(horizontal="center" if cell.column <= 7 else "left", vertical="center", wrap_text=True)
            cell.border = Border(left=thin_gray, right=thin_gray, top=blue_line, bottom=blue_line)
            if cell.column <= 3:
                cell.fill = light_fill
            elif cell.row % 2 == 0:
                cell.fill = stripe_fill

    for col, width in {"A": 24, "B": 18, "C": 30, "D": 10, "E": 12, "F": 12, "G": 9, "H": 48, "I": 12}.items():
        ws.column_dimensions[col].width = width
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 34
    for col in ["E", "F"]:
        for row in range(2, len(rows) + 2):
            ws[f"{col}{row}"].number_format = "yyyy-mm-dd"
    merge_runs(ws, 1, 2, [r[0] for r in rows])
    merge_runs(ws, 2, 2, [(r[0], r[1]) for r in rows])
    ws.freeze_panes = "D2"

    timeline = wb.create_sheet("일정표")
    timeline.sheet_view.showGridLines = False
    first = min(d(r[4]) for r in rows).replace(day=1)
    last_raw = max(d(r[5]) for r in rows)
    last = date(last_raw.year + (last_raw.month // 12), (last_raw.month % 12) + 1, 1) - timedelta(days=1)
    dates = []
    cursor = first
    while cursor <= last:
        dates.append(cursor)
        cursor += timedelta(days=1)

    last_col = 4 + len(dates) - 1
    last_col_letter = get_column_letter(last_col)
    timeline.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    timeline["A2"] = "■ WBS 작업 일정"
    timeline["A2"].font = Font(name="맑은 고딕", size=12)
    timeline["A2"].alignment = Alignment(horizontal="left", vertical="center")
    timeline.merge_cells("A3:B4")
    timeline["A3"] = "일정계획"
    timeline.merge_cells("C3:C4")
    timeline["C3"] = "3Depth 작업"

    for cell_ref in ["A3", "C3"]:
        cell = timeline[cell_ref]
        cell.fill = PatternFill("solid", fgColor="EDEDED")
        cell.font = Font(name="맑은 고딕", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    month_start = 0
    for idx, day in enumerate(dates + [None]):
        if idx == len(dates) or day.month != dates[month_start].month:
            start_col = 4 + month_start
            end_col = 4 + idx - 1
            timeline.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
            cell = timeline.cell(3, start_col)
            cell.value = f"{dates[month_start].month}월"
            cell.fill = PatternFill("solid", fgColor="EDEDED")
            cell.font = Font(name="맑은 고딕", bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            month_start = idx

    for idx, day in enumerate(dates):
        cell = timeline.cell(4, 4 + idx)
        cell.value = day
        cell.number_format = "d"
        cell.fill = PatternFill("solid", fgColor="EDEDED")
        cell.font = Font(name="맑은 고딕", bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 5
    for idx, item in enumerate(rows):
        row = start_row + idx
        timeline.cell(row, 1).value = item[0]
        timeline.cell(row, 2).value = item[1]
        timeline.cell(row, 3).value = item[2]
        for col in range(1, last_col + 1):
            cell = timeline.cell(row, col)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal="center" if col <= 2 else "left", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="D9D9D9" if col >= 4 and dates[col - 4].weekday() >= 5 else "FFFFFF")

    merge_runs(timeline, 1, start_row, [r[0] for r in rows])
    merge_runs(timeline, 2, start_row, [(r[0], r[1]) for r in rows])

    thin = Side(style="thin", color="000000")
    date_grid = Side(style="thin", color="B7B7B7")
    month_grid = Side(style="medium", color="000000")
    month_start_cols = {4}
    for idx in range(1, len(dates)):
        if dates[idx].month != dates[idx - 1].month:
            month_start_cols.add(4 + idx)
    for row in range(3, start_row + len(rows)):
        for col in range(1, last_col + 1):
            cell = timeline.cell(row, col)
            if col >= 4:
                left = month_grid if col in month_start_cols else date_grid
                cell.border = Border(left=left, right=date_grid, top=thin, bottom=thin)
            else:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    bar_range = f"D{start_row}:{last_col_letter}{start_row + len(rows) - 1}"
    for status, color in [("미진행", "FFC000"), ("진행중", "2F5597"), ("검토중", "92D050"), ("완료", "A6A6A6")]:
        timeline.conditional_formatting.add(
            bar_range,
            FormulaRule(
                formula=[f'=AND(D$4>=WBS목록!$E2,D$4<=WBS목록!$F2,WBS목록!$D2="{status}")'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )
    timeline.column_dimensions["A"].width = 24
    timeline.column_dimensions["B"].width = 18
    timeline.column_dimensions["C"].width = 31
    for col in range(4, last_col + 1):
        timeline.column_dimensions[get_column_letter(col)].width = 3
    timeline.freeze_panes = "D5"
    wb.save(OUTPUT)
    wb.save(ARCHIVE)


def write_json() -> None:
    DATA.write_text(
        json.dumps(
            [
                {
                    "depth1": r[0],
                    "depth2": r[1],
                    "depth3": r[2],
                    "status": r[3],
                    "start": r[4],
                    "end": r[5],
                    "next_action": r[6],
                    "notion": r[7],
                    "repo": "https://github.com/sehuniotech-ctrl/Iotech_sems" if r[0] == "아이오테크 SEMS SL 개발" else "",
                }
                for r in rows
            ],
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    build_workbook()
    write_json()
    print(OUTPUT)
